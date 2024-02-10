import openpyxl


class ClearDataPlanning:
    def __init__(self, file_path, sheet_name, start_row, start_column):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.start_row = start_row
        self.start_column = start_column

    def clear_data_in_range(
        self,
        end_row,
        end_column,
    ):
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook[self.sheet_name]

        # If end_row or end_column is not specified, use the maximum available rows/columns
        if end_row is None:
            end_row = sheet.max_row
        if end_column is None:
            end_column = sheet.max_column

        # Clear the data in the specified range
        for row in sheet.iter_rows(
            min_row=self.start_row,
            max_row=end_row,
            min_col=self.start_column,
            max_col=end_column,
        ):
            for cell in row:
                cell.value = None

        # Save the modified workbook
        workbook.save(self.file_path)
        print(
            f"Data cleared from {self.sheet_name}! Range: {self.start_row}:{self.start_column} - {end_row}:{end_column}"
        )
        return True
