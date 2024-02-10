import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
import string
import datetime
import warnings
import json
import ast
from DataPlan import DataPlan

# ปิดการแสดงข้อความแจ้งเตือนทั้งหมด
warnings.filterwarnings("ignore")


class PlanningProcess:
    def __init__(
        self,
        file_path,
        sheet_name,
        status_plan,
        prd_start_date,
        current_column,
        current_value,
        current_row,
        ot_plan,
        targetModel,
    ):
        self.plan = openpyxl.load_workbook(file_path)
        self.sheet = self.plan[sheet_name]
        self.file_path = file_path
        self.status_plan = status_plan
        self.prd_start_date = prd_start_date
        self.current_column = current_column
        self.current_value = current_value
        self.current_row = current_row
        self.ot_plan = ot_plan

        self.data_json = {}
        with open("parameter.json", "r") as json_file:
            self.data_json = json.load(json_file)
        self.column_line = self.data_json["column_line"]  # U
        self.column_index = self.data_json["column_index"]  # เริ่มต้นที่ Column "V"
        self.date_row = self.data_json["date_row"]
        self.start_row = self.data_json["start_row"]
        self.batch_column = self.data_json["batch_column"]  # B
        self.type_column = self.data_json["type_column"]  # C
        self.model_column = self.data_json["model_column"]  # E
        self.qty_column = self.data_json["qty_column"]  # F
        self.eta_column = self.data_json["eta_column"]  # N
        self.ot_batch = self.data_json["ot_batch"]  # BATCH HAVE TO OPEN OT
        self.holiday = [
            datetime.datetime.strptime(holiday, "%Y-%m-%d")
            for holiday in self.data_json["holiday"]
        ]

        # find last row and get A4 to C at Last row
        self.last_row = self.sheet.max_row
        end_cell = f"F{self.last_row}"
        end_cell_line = f"R{self.last_row}"
        self.datas = list(self.sheet[f"B11:{end_cell}"])  # Include columns B to F
        self.lineCol = self.sheet["R11":end_cell_line]
        self.targetModel = targetModel
        self.remain = []

    def getValues(self, arry):
        block = []
        for row in arry:
            for cell in row:
                if cell.value is not None:
                    block.append(cell.value)
        return block

    def getColName(self):
        num_columns = self.sheet.max_column - self.column_index
        col_exel = []
        for i in range(num_columns):
            col_number = i + self.column_index
            col_name = ""
            while col_number >= 0:
                col_name = string.ascii_uppercase[col_number % 26] + col_name
                col_number = col_number // 26 - 1
            col_exel.append(col_name)
        return col_exel

    def reshape_data(self, data):
        reshaped_data = []
        for i in range(0, len(data), 5):
            batch = data[i : i + 5]
            reshaped_data.append(batch)
        return reshaped_data

    def paramChangeModel(self, prv_model, model):
        if model == prv_model:
            return 0
        else:
            return 1

    def getDayOfWeek(self, date_string):
        try:
            if self.is_datetime(date_string):
                day_of_week = date_string.strftime("%A")
            else:
                date_object = datetime.datetime.strptime(date_string, "%d/%m/%Y")
                day_of_week = date_object.strftime("%A")
            return day_of_week
        except ValueError:
            return "Invalid date format"

    def getDay(self, array, row_start):
        block = []
        col = 0
        while col < len(array):
            if (
                self.getDayOfWeek(self.sheet[f"{array[col]}{row_start}"].value)
                == "Invalid date format"
            ):
                break
            block.append(
                self.getDayOfWeek(self.sheet[f"{array[col]}{row_start}"].value)
            )
            col += 1
        return block

    def getFirstPRD(self, data):
        unique_values = set(item[-1] for item in data)
        result = {}
        i = data[0][-1]

        for value in unique_values:
            for item in data:
                if item[-1] == value:
                    result[i] = item[4]
                    break
            i += 1
        return result

    def getColLine(self, col_names, lines, row_start):
        line_set = {}
        for index in lines:
            i = 0
            for col_name in col_names:
                cell = self.sheet[f"{col_name}{row_start}"]
                if (
                    isinstance(cell.value, datetime.datetime)
                    and cell.value == lines[index]
                ):
                    line_set[index] = col_names[i:]
                    break
                i += 1
        return line_set

    def getColLineContinue(self, lines, col_names):
        line_set = {}
        current_column = self.current_column
        print("Data getColLineContinue:", lines, current_column)
        for line, col in zip(lines, current_column):
            i = 0
            for col_name in col_names:
                if col_name == col:
                    line_set[line] = col_names[i:]
                    break
                i += 1
        return line_set

    def add_next_excel_column(self, column_list):
        if not column_list:
            return ["A"]  # เริ่มต้นที่ 'A' ถ้า list ว่าง

        last_column = column_list[-1]
        next_column = list(last_column)  # เปลี่ยนเป็น list ของตัวอักษร

        # วนลูปจากท้าย list
        i = len(next_column) - 1
        while i >= 0:
            if next_column[i] == "Z":
                next_column[i] = "A"
            else:
                next_column[i] = chr(ord(next_column[i]) + 1)
                break
            i -= 1

        # ถ้าทุกตัวเป็น 'Z', เพิ่ม 'A' ที่หน้า
        if i < 0:
            next_column.insert(0, "A")

        return column_list + ["".join(next_column)]

    def add_date(self, date_column):
        cell = self.sheet[f"{date_column}{self.date_row}"]
        if cell.value:
            # แปลงค่าใน cell เป็น datetime
            if isinstance(cell.value, datetime.datetime):
                # ค่าเป็น datetime.datetime แล้ว
                date_value = cell.value
            else:
                # ค่าเป็น string, แปลงเป็น datetime
                date_value = datetime.datetime.strptime(cell.value, "%d/%m/%Y")
            # เพิ่ม 1 วัน
            new_date_value = date_value + datetime.timedelta(days=1)
        return new_date_value.strftime("%d/%m/%Y")

    def get_prd_line(self, data_list):
        unique_items = {}
        for item in data_list:
            last_element = item[-1]
            if last_element not in unique_items:
                unique_items[last_element] = last_element
        unique_result = list(unique_items.values())
        return unique_result

    def border_cell(self, row):
        last_column = self.sheet.max_column
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

        for col_num in range(1, last_column + 1):
            cell = self.sheet.cell(row=row, column=col_num)
            cell_date = self.sheet.cell(row=self.date_row, column=col_num)
            cell.border = border

            # ตรวจสอบว่าค่าในเซลล์เป็น datetime object หรือไม่
            if self.is_datetime(cell_date.value) and col_num > 20:
                if self.getDayOfWeek(cell_date.value) == "Sunday":
                    cell.fill = fill

    def is_datetime(self, value):
        return isinstance(value, datetime.datetime)

    def add_next_column(self, columns):
        if not columns:
            return ["A"]
        last_column = columns[-1]
        next_column = self.increment_column(last_column)
        columns.append(next_column)
        return columns

    def increment_column(self, column):
        num_value = 0
        for char in column:
            num_value = num_value * 26 + (ord(char) - ord("A") + 1)
        num_value += 1
        result_column = ""
        while num_value > 0:
            num_value, remainder = divmod(num_value - 1, 26)
            result_column = chr(65 + remainder) + result_column
        return result_column

    def insert_row(self, row):
        row = row + 1
        self.sheet.insert_rows(row, amount=1)
        # self.sheet.cell(row=row, column=2, value="")
        col_list = [7, 8, 9, 11, 14]
        val_list = [
            f"=SUM(S{row}:LF{row})",
            f"=G{row}-F{row}",
            f"=XLOOKUP(XMATCH(FALSE,ISBLANK(S{row}:LE{row})),$S$9:$LE$9,$S$10:$LE$10)",
            f"=I{row}-M{row}",
            f"=XLOOKUP(E{row},[RecoveredExternalLink2]Sheet1!$D$2:$D$3720,[RecoveredExternalLink2]Sheet1!$A$2:$A$3720,0)",
        ]
        for col_num, value in zip(col_list, val_list):
            self.sheet.cell(row=row, column=col_num, value=value)
            self.border_cell(row)
        self.plan.save(self.file_path)

    def get_last_row(self):
        last_row_with_data = 1
        for row in range(1, self.sheet.max_row + 1):
            # ถ้าเซลล์มีค่า, อัปเดต last_row_with_data
            if self.sheet[f"B{row}"].value is not None:
                last_row_with_data = row
            else:
                # ถ้าเจอเซลล์ที่ไม่มีข้อมูล, หยุดการวนลูป
                break
        return last_row_with_data

    def putTarget(self, data, map_man_power):
        col_names = self.getColName()
        first_prd_lines = self.prd_start_date
        # เซ็ตค่าตัวเเปร
        last_row = self.get_last_row()
        rowStart = (
            self.start_row if self.status_plan == "NEW SEASON" else self.current_row
        )
        ot = 0
        col = 0
        row = 0
        uph = 0
        i = 0
        index = 0
        qty = 0
        actual_per_day = []
        obj_plan = []
        planning_data = {}
        prd_batch = []
        prd_date = []
        count = 0
        cmt_count = 0
        prd_day = []
        sequense = []
        msg = []
        lines = self.get_prd_line(data)
        map_line = {1: "H1", 2: "H2", 3: "H3", 4: "H4"}
        col_line = (
            self.getColLine(col_names, first_prd_lines, self.date_row)
            if self.status_plan == "NEW SEASON"
            else self.getColLineContinue(lines, col_names)
        )
        for line in lines:
            col_exel = col_line[line]
            rowStart = (
                rowStart
                if self.status_plan == "NEW SEASON"
                else self.current_row[index]
            )
            man_power = map_man_power[line]
            while row < len(data) and data[row][-1] == line:
                qty = int(data[row][3])
                while int(data[row][3]) > 0:
                    date_check = self.getDay(col_exel, self.date_row)
                    # prd parameter
                    uph = float(self.targetModel[data[row][2]]) * man_power
                    working_time = 8
                    # ----- add one column Date -----
                    if col == len(date_check):
                        col_exel = self.add_next_excel_column(col_exel)
                        self.sheet[f"{col_exel[col]}{self.date_row}"] = self.add_date(
                            col_exel[col - 1]
                        )
                        self.sheet[f"{col_exel[col]}{self.date_row}"].number_format = (
                            "DD/MM/YYYY"
                        )
                        self.sheet[f"{col_exel[col]}{self.date_row}"].font = Font(
                            bold=True
                        )
                        self.sheet[f"{col_exel[col]}{self.date_row}"].alignment = (
                            Alignment(horizontal="center", vertical="center")
                        )
                        date_check = self.getDay(col_exel, self.date_row)
                    if data[row][0] in self.ot_batch:
                        working_time = 10
                        ot += 1
                    elif date_check[col] in self.ot_plan:
                        working_time = 10
                    upd = round((uph * working_time), 0)
                    if col >= len(col_exel):
                        col_exel = self.add_next_column(col_exel)
                        self.sheet[f"{col_exel[col]}{self.date_row}"] = self.sheet[
                            f"{col_exel[col-1]}{self.date_row}"
                        ].value + datetime.timedelta(days=1)
                        date_check.append(
                            self.getDayOfWeek(
                                self.sheet[f"{col_exel[col]}{self.date_row}"].value
                            )
                        )
                    # today is Sunday ?
                    if (
                        date_check[col] == "Sunday"
                        or self.sheet[f"{col_exel[col]}{self.date_row}"].value
                        in self.holiday
                    ):
                        # shift 1 col
                        pass
                    elif (
                        int(data[row][3]) == qty
                        and len(self.current_column) > 0
                        and self.current_value[index] > upd
                        and self.current_value[index] > 0
                    ):
                        actual_per_day.append(round(self.current_value[index], 0))
                        self.current_value[index] = 0
                    elif (
                        int(data[row][3]) == qty
                        and len(self.current_column) > 0
                        and self.current_value[index] < upd
                        and self.current_value[index] > 0
                    ):
                        cmt_factor = self.paramChangeModel(
                            data[row - 1][1], data[row][1]
                        )
                        remainChangeBatch = round(
                            (upd - self.current_value[index]) - (cmt_factor * uph),
                            0,
                        )
                        self.sheet[f"{col_exel[col]}{rowStart}"] = round(
                            remainChangeBatch, 0
                        )
                        actual_per_day.append(
                            round(abs(remainChangeBatch), 0)
                        )  # Actual
                        prd_batch.append(abs(remainChangeBatch))  # Actual
                        prd_date.append(
                            self.sheet[f"{col_exel[col]}{self.date_row}"].value
                        )
                        data[row][3] -= remainChangeBatch
                        sequense.append(f"{remainChangeBatch}:B0")
                        self.current_value[index] = 0
                    elif (row == 0 and int(data[row][3]) >= upd) or (
                        row != 0
                        and data[row - 1][-1] != data[row][-1]
                        and int(data[row][3]) >= upd
                    ):
                        data[row][3] -= upd
                        self.remain.append(int(data[row][3]))
                        actual_per_day.append(upd)  # Actual
                        prd_batch.append(upd)  # Actual
                        self.sheet[f"{col_exel[col]}{rowStart}"] = round(abs(upd), 0)
                        prd_date.append(
                            self.sheet[f"{col_exel[col]}{self.date_row}"].value
                        )
                        count += 1
                        sequense.append(f"{upd}:B1")
                    # commond batch
                    elif (
                        len(self.remain) > 0
                        and self.remain[-1] < upd
                        and int(data[row][3]) > upd
                    ):
                        # get previous upd
                        prv_uph = float(self.targetModel[data[row - 1][2]]) * man_power
                        prv_upd = prv_uph * working_time
                        cmt_factor = self.paramChangeModel(
                            data[row - 1][1], data[row][1]
                        )
                        # count change model group by cabinet
                        if cmt_factor == 1:
                            cmt_count += 1
                        # check scarp
                        if actual_per_day[-1] != prv_upd and upd > actual_per_day[-1]:
                            # check joint change batch
                            if (
                                (upd - (cmt_factor * uph)) - actual_per_day[-1] <= upd
                                or (upd - (cmt_factor * uph)) - actual_per_day[-1]
                                <= prv_upd
                            ) and date_check[col] != "Monday":
                                col = col - 1
                            # shift 2 column for monday to sat
                            elif (
                                (upd - (cmt_factor * uph)) - actual_per_day[-1] <= upd
                                or (upd - (cmt_factor * uph)) - actual_per_day[-1]
                                <= prv_upd
                            ) and date_check[col] == "Monday":
                                col = col - 2
                            while (
                                self.sheet[f"{col_exel[col]}{self.date_row}"].value
                                in self.holiday
                                or date_check[col] == "Sunday"
                            ):
                                col = col - 1
                            if date_check[col] in self.ot_plan:
                                working_time = 10
                                ot += 1
                            else:
                                working_time = 8
                            upd = round((uph * working_time), 0)
                            prv_upd = (
                                float(self.targetModel[data[row - 1][2]])
                                * man_power
                                * working_time
                            )
                            remain_hour_prv = abs(
                                round(prv_upd - (cmt_factor * prv_uph), 0) / prv_uph
                            ) - (round(actual_per_day[-1], 0) / prv_uph)
                            remainChangeBatch = round(uph * remain_hour_prv, 0)
                            actual_per_day.append(remainChangeBatch)
                            prd_batch.append(remainChangeBatch)
                            if remainChangeBatch > 0:
                                self.sheet[f"{col_exel[col]}{rowStart}"] = round(
                                    remainChangeBatch, 0
                                )
                                count += round(remainChangeBatch / upd, 0)
                                sequense.append(f"{remainChangeBatch}C1")
                            data[row][3] -= remainChangeBatch
                            self.remain.append(data[row][3])
                            prd_date.append(
                                self.sheet[f"{col_exel[col]}{self.date_row}"].value
                            )
                        else:
                            data[row][3] -= upd
                            self.sheet[f"{col_exel[col]}{rowStart}"] = round(
                                abs(upd), 0
                            )
                            actual_per_day.append(round(upd, 0))  # Actual
                            prd_batch.append(upd)  # Actual
                            prd_date.append(
                                self.sheet[f"{col_exel[col]}{self.date_row}"].value
                            )
                            self.remain.append(data[row][3])
                            count += 1
                            sequense.append(f"{upd}C2")
                    # input qty if qty <= upd
                    elif int(data[row][3]) <= upd:
                        if (
                            len(self.remain) > 0
                            and len(actual_per_day) > 0
                            and self.remain[-1] == 0
                            and actual_per_day[-1] < upd
                            and (int(data[row][3]) + actual_per_day[-1]) <= upd
                        ):
                            cmt_factor = self.paramChangeModel(
                                data[row - 1][1], data[row][1]
                            )
                            if cmt_factor == 1:
                                cmt_count += 1
                            col = col - 1 if date_check[col] != "Monday" else col - 2
                            while (
                                self.sheet[f"{col_exel[col]}{self.date_row}"].value
                                in self.holiday
                                or date_check[col] == "Sunday"
                            ):
                                col = col - 1
                            if date_check[col] in self.ot_plan:
                                working_time = 10
                                ot += 1
                            else:
                                working_time = 8
                            upd = round((uph * working_time), 0)
                            remainChangeBatch = round(
                                ((upd - (cmt_factor * uph)) - actual_per_day[-1]),
                                0,
                            )
                            self.sheet[f"{col_exel[col]}{rowStart}"] = int(data[row][3])
                            actual_per_day.append(int(data[row][3]))  # Actual
                            prd_batch.append(int(data[row][3]))  # Actual
                            prd_date.append(
                                self.sheet[f"{col_exel[col]}{self.date_row}"].value
                            )
                            sequense.append(f"{int(data[row][3])}D1")
                            data[row][3] -= int(data[row][3])
                            count += int(data[row][3]) / upd
                            self.remain.append(int(data[row][3]))

                        else:
                            self.sheet[f"{col_exel[col]}{rowStart}"] = round(
                                abs(int(data[row][3])), 0
                            )
                            actual_per_day.append(round(int(data[row][3]), 0))  # Actual
                            prd_batch.append(int(data[row][3]))  # Actual
                            prd_date.append(
                                self.sheet[f"{col_exel[col]}{self.date_row}"].value
                            )
                            sequense.append(f"{data[row][3]}D2")
                            data[row][3] = 0
                            count += int(data[row][3]) / upd
                            self.remain.append(0)

                    # check prev value  and col_exel[col] == sequense[-1]
                    elif int(data[row][3]) >= upd:
                        data[row][3] -= upd
                        self.sheet[f"{col_exel[col]}{rowStart}"] = round(abs(upd), 0)
                        actual_per_day.append(round(upd, 0))  # Actual
                        prd_batch.append(upd)  # Actual
                        prd_date.append(
                            self.sheet[f"{col_exel[col]}{self.date_row}"].value
                        )
                        self.remain.append(data[row][3])
                        count += 1
                        sequense.append(f"{upd}E")

                    # nomal put
                    else:
                        self.sheet[f"{col_exel[col]}{rowStart}"] = round(
                            abs(int(data[row][3])), 0
                        )
                        actual_per_day.append(round(int(data[row][3]), 0))  # Actual
                        sequense.append(f"{int(data[row][3])}F")
                        data[row][3] = 0
                        prd_batch.append(upd)  # Actual
                        prd_date.append(
                            self.sheet[f"{col_exel[col]}{self.date_row}"].value
                        )
                        count += 1
                        self.remain.append(0)
                    msg = [True]
                    col += 1

                # ["B" = BATCH, "C" = TYPE, "E" = MODEL, "F" = QTY, "N" = ETA]
                self.sheet[f"{self.batch_column}{rowStart}"] = data[row][0]
                self.sheet[f"{self.type_column}{rowStart}"] = data[row][1]
                self.sheet[f"{self.model_column}{rowStart}"] = data[row][2]
                self.sheet[f"{self.qty_column}{rowStart}"] = qty
                self.sheet[f"{self.eta_column}{rowStart}"] = data[row][-3]
                self.sheet[f"{self.column_line}{rowStart}"] = map_line[data[row][-1]]
                if rowStart > last_row:
                    self.border_cell(rowStart)

                # check next data
                row_next = row + 1 if row + 1 <= len(data) - 1 else len(data) - 1
                line_next_row = self.sheet[f"{self.column_line}{rowStart + 1}"].value
                if (
                    data[row_next][-1] == line
                    and line_next_row != map_line[line]
                    and line_next_row is not None
                ):
                    if index < len(self.current_row):
                        for i in range(len(self.current_row)):
                            self.current_row[i] = self.current_row[i] + 1
                        self.insert_row(rowStart)
                # GET DATA
                # obj_plan = [data[row]]
                # print("PRD DATE:", prd_date)
                obj_plan.append(prd_date)  # DATE_PRD
                obj_plan.append(prd_batch)  # ACTUAL_PRD
                batch = [data[row]][0][0]
                planning_data[str(batch)] = obj_plan
                obj_plan = []
                prd_batch = []
                prd_date = []
                prd_day.append(round(count, 2))
                row += 1
                rowStart += 1
                count = 0
                i += 1
            # print("Date Check:", date_check)
            # print(f"Logic: {sequense}")
            # print(f"Day Batch: {prd_day}")
            # print(f"Change Model: {cmt_count}")
            # print(f"Remain: {self.remain}")
            # print(f"Daily Actual: {actual_per_day}")
            # reward = k * (1 - (number_of_model_changes / total_production_days))
            # print([prd_day, [cmt_count], actual_per_day, msg])
            # print("PRD DATE:", prd_date)
            # print(planning_data)
            self.plan.save(self.file_path)
            sequense = []
            actual_per_day = []
            col = 0
            index += 1
        ot = ot
        return planning_data

    def return_data(self, data, line):
        i = 0
        for arry in data:
            arry.append(line[i])
            del arry[2]
            i += 1
        return data
