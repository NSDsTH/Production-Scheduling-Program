import openpyxl
import ast
import warnings
import string
import datetime
import time
import json
from ClearData import ClearDataPlanning

warnings.simplefilter("ignore")


class DataPlan:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.column_letters = ["B", "C", "E", "F", "N"]
        self.data_json = {}
        self.prd_start_date = {}
        with open("parameter.json", "r") as json_file:
            self.data_json = json.load(json_file)
            self.prd_start_date = {
                1: datetime.datetime.strptime(
                    self.data_json["prd_start_h1"], "%Y-%m-%d"
                ),
                2: datetime.datetime.strptime(
                    self.data_json["prd_start_h2"], "%Y-%m-%d"
                ),
                3: datetime.datetime.strptime(
                    self.data_json["prd_start_h3"], "%Y-%m-%d"
                ),
                4: datetime.datetime.strptime(
                    self.data_json["prd_start_h4"], "%Y-%m-%d"
                ),
            }
        self.eta_column = self.data_json["eta_column"]  # U
        self.column_line = self.data_json["column_line"]  # U
        self.column_index = self.data_json["column_index"]  # V OR 21
        self.start_row = self.data_json["start_row"]  # 12
        self.prd_line = self.data_json["prd_line"]
        self.date_row = self.data_json["date_row"]
        self.map_line = {
            "H1": "HISENSE 1",
            "H2": "HISENSE 2",
            "H3": "HISENSE 3",
            "H4": "HISENSE 4",
        }

    def is_datetime(self, value):
        return isinstance(value, datetime.datetime)

    # def is_datetime_excel(self, value):
    #     try:
    #         datetime_object = datetime.datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
    #         return True
    #     except ValueError:
    #         return False

    def get_data_from_sheet(self):
        try:
            plan = openpyxl.load_workbook(self.file_path, data_only=True)
            sheet = plan[self.sheet_name]
            start_row = self.start_row
            last_row = sheet.max_row
            data = []
            row = []

            while start_row <= last_row:
                for col in self.column_letters:
                    cell_value = sheet[f"{col}{start_row}"]
                    row.append(cell_value.value)

                if self.is_datetime(row[-1]):
                    data.append(row)
                row = []
                start_row += 1

            return data
        except Exception as e:
            print("เกิดข้อผิดพลาดในการอ่านไฟล์ Excel:", e)
            return None

    def getColName(self):
        plan = openpyxl.load_workbook(self.file_path, data_only=True)
        sheet = plan[self.sheet_name]
        column_index = self.column_index
        num_columns = sheet.max_column - column_index
        col_exel = []
        for i in range(num_columns):
            col_number = i + column_index
            col_name = ""
            while col_number >= 0:
                col_name = string.ascii_uppercase[col_number % 26] + col_name
                col_number = col_number // 26 - 1
            col_exel.append(col_name)
        return col_exel

    def getCurrentColumn(self):
        plan = openpyxl.load_workbook(self.file_path, data_only=True)
        sheet = plan[self.sheet_name]
        map_line = {"HISENSE 1": 1, "HISENSE 2": 2, "HISENSE 3": 3, "HISENSE 4": 4}
        prd_line = [map_line[line] for line in self.prd_line]
        current_date = self.prd_start_date[prd_line[0]].date()
        columnName = self.getColName()
        num_col = self.column_index + 1
        current_column = None
        for column_index in columnName:  # เริ่มต้นที่คอลัมน์ V (คอลัมน์ 22)
            cell = sheet[f"{column_index}{self.date_row}"]
            if (
                isinstance(cell.value, datetime.datetime)
                and cell.value.date() == current_date
            ):
                current_column = column_index
                break
            num_col += 1

        # ค่าความปลอดภัยในการเเก้ไขการผลิต ควรเป็น 3 batch
        shift_batch = 0
        num_col = num_col + shift_batch
        return num_col

    def getRealList(self, my_list):
        result = []
        for i in range(len(my_list)):
            if i < len(my_list) - 1:
                if abs(my_list[i] - my_list[i + 1]) in [1, 2, 3]:
                    pass
                else:
                    result.append(my_list[i] + 1)
            else:
                if abs(my_list[i] - my_list[i - 1]) in [1, 2, 3]:
                    result.append(my_list[i] + 1)
                else:
                    result.append(my_list[i] + 1)
        return result

    def getRemainBatch(self):
        plan = openpyxl.load_workbook(self.file_path, data_only=True)
        sheet = plan[self.sheet_name]
        num_col = self.getCurrentColumn()
        row_start = self.start_row
        list_current_row = []
        current_row = []
        row_empty_line = []
        list_end_row = []
        batchs = []
        try:
            for row in sheet.iter_rows(
                min_row=self.start_row, min_col=num_col, max_col=num_col
            ):
                for cell in row:
                    line = (
                        self.map_line[sheet[f"{self.column_line}{row_start}"].value]
                        if sheet[f"{self.column_line}{row_start}"].value is not None
                        else None
                    )
                    eta = (
                        sheet[f"{self.eta_column}{row_start}"].value
                        if sheet[f"{self.eta_column}{row_start}"].value is not None
                        else None
                    )
                    # print("Line:", line)
                    # print("cell:", cell.value)
                    # print("Row:", row_start)
                    # print("ETA:", eta)
                    if cell.value and line in self.prd_line:
                        current_row.append(row_start)
                    if self.is_datetime(eta) and line == None:
                        row_empty_line.append(row_start)
                row_start += 1
            list_current_row = self.getRealList(current_row)
            for item in list_current_row:
                rows = item
                while True:
                    test = [
                        sheet[f"{column}{rows}"].value for column in self.column_letters
                    ]
                    if (
                        sheet[f"{self.column_line}{item}"].value
                        == sheet[f"{self.column_line}{rows}"].value
                    ):
                        batch_check = [
                            sheet[f"{column}{rows}"].value
                            for column in self.column_letters
                        ]
                        if None not in batch_check and self.is_datetime(
                            batch_check[-1]
                        ):
                            batchs.append(batch_check)
                        rows += 1
                    else:
                        clear_data = ClearDataPlanning(
                            self.file_path,
                            self.sheet_name,
                            start_row=item,  # row เเรกในการล้าง
                            start_column=num_col,
                        )
                        clear_data.clear_data_in_range(rows - 1, None)
                        list_end_row.append(rows - 1)
                        break
                for row_empty in row_empty_line:
                    batch_empty = [
                        sheet[f"{column}{row_empty}"].value
                        for column in self.column_letters
                    ]
                    batchs.append(batch_empty)
                if len(row_empty_line) > 0:
                    # ล้าง row ของ batch ที่ยังไม่ได้วางเเผน
                    clear_data = ClearDataPlanning(
                        self.file_path,
                        self.sheet_name,
                        start_row=row_empty_line[0],
                        start_column=0,
                    )
                    clear_data.clear_data_in_range(row_empty_line[-1], None)
            return batchs, list_current_row
        except Exception as e:
            print(f"เกิดข้อผิดพลาด: {str(e)}")
            return None, None  # Return None instead of False

    def find_last_value_column_in_row(self, target_list):
        try:
            with open("parameter.json", "r") as json_file:
                data = json.load(json_file)
            plan = openpyxl.load_workbook(self.file_path, data_only=True)
            sheet = plan[self.sheet_name]
            map_line = {
                "HISENSE 1": "prd_start_h1",
                "HISENSE 2": "prd_start_h2",
                "HISENSE 3": "prd_start_h3",
                "HISENSE 4": "prd_start_h4",
            }
            map_line_prd = {
                "prd_start_h1": 1,
                "prd_start_h2": 2,
                "prd_start_h3": 3,
                "prd_start_h4": 4,
            }
            lines = [map_line[line] for line in self.prd_line]
            list_column = []
            list_value = []
            list_start_date = {}
            if len(target_list) > 0:
                for target_row, line in zip(target_list, lines):
                    last_column = None
                    target_row = target_row - 1  # Previous row
                    for cell in sheet.iter_rows(min_row=target_row, max_row=target_row):
                        for cell in cell:
                            if cell.value:
                                last_column = cell.column_letter
                    if last_column:
                        list_value.append(sheet[f"{last_column}{target_row}"].value)
                        list_column.append(last_column)
                    date_obj = sheet[f"{last_column}{self.date_row}"].value.date()
                    formatted_date = date_obj.strftime("%Y-%m-%d")
                    data[line] = formatted_date
                    list_start_date[map_line_prd[line]] = datetime.datetime.strptime(
                        formatted_date, "%Y-%m-%d"
                    )
            with open("parameter.json", "w") as json_file:
                json.dump(data, json_file)
            print("target_list:", target_list)
            print("list_column:", list_column)
            print("list_value:", list_value)
            print("Start Date:", list_start_date)
            return list_column, list_value, list_start_date
        except Exception as e:
            print(f"เกิดข้อผิดพลาด: เกิน Column ที่มีอยู่")
            print(f"เกิดข้อผิดพลาด: {str(e)}")
            return None, None


# if __name__ == "__main__":
#     file_path = "Production plan.xlsx"
#     prd_sheet_name = "Production plan"
#     data_plan = DataPlan(file_path, prd_sheet_name)
#     schedule_plan, list_current_row = data_plan.getRemainBatch()
#     list_col, list_val, list_start_date = data_plan.find_last_value_column_in_row(
#         list_current_row
#     )
#     for batch in schedule_plan:
#         print(batch)
